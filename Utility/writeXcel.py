
import openpyxl
import datetime


class xlrWrite:
    def writecpu(self,cpu):
        wb = openpyxl.load_workbook('Utility/machineReport.xlsx')
        ws = wb['Sheet1']
        rows = ws.max_row
        ws.cell(row=rows, column=2).value = cpu
        wb.save("Utility/machineReport.xlsx")


    def writememory(self,memory):
        wb = openpyxl.load_workbook('Utility/machineReport.xlsx')
        ws = wb['Sheet1']
        rows = ws.max_row
        ws.cell(row=rows + 1, column=1).value = str(datetime.datetime.now())
        ws.cell(row=rows + 1, column=3).value = memory
        wb.save("Utility/machineReport.xlsx")

    def writeusage(self,usage):
        wb = openpyxl.load_workbook('Utility/machineReport.xlsx')
        ws = wb['Sheet1']
        rows = ws.max_row
        ws.cell(row=rows, column=4).value = usage
        wb.save("Utility/machineReport.xlsx")
